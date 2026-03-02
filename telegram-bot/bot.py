import os
import io
import json
import logging
import httpx
import gspread
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, filters, ContextTypes,
)
from telegram.constants import ParseMode

# ─── Настройки ───────────────────────────────────────────────────────────────

logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger("barbercrm_bot")

BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
BOT_PASSWORD = os.getenv("TELEGRAM_BOT_PASSWORD", "admin")
BACKEND_URL = os.getenv("BACKEND_URL", "http://backend:8000")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]

# Авторизованные chat_id (хранятся в памяти, сбрасываются при рестарте)
authorized_users: set[int] = set()

# ─── Google Sheets клиент (для получения списка филиалов) ────────────────────

_sheets_client = None

def get_sheets_client():
    global _sheets_client
    if _sheets_client:
        return _sheets_client
    try:
        sa_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "{}")
        info = json.loads(sa_json)
        if "private_key" in info:
            info["private_key"] = info["private_key"].replace("\\n", "\n")
        creds = Credentials.from_service_account_info(info, scopes=SCOPES)
        _sheets_client = gspread.authorize(creds)
        return _sheets_client
    except Exception as e:
        logger.error(f"Ошибка Google Sheets: {e}")
        return None


def get_branches_list() -> list[dict]:
    """Получить список филиалов из главной таблицы."""
    try:
        client = get_sheets_client()
        if not client:
            return []
        sheet_id = os.getenv("GOOGLE_SHEET_ID", "")
        spreadsheet = client.open_by_key(sheet_id)
        ws = spreadsheet.worksheet("Филиалы")
        records = ws.get_all_records()
        return [
            {
                "name": r.get("Название", ""),
                "address": r.get("Адрес", ""),
                "manager": r.get("Имя руководителя", ""),
                "phone": r.get("Телефон", ""),
            }
            for r in records
            if r.get("Название")
        ]
    except Exception as e:
        logger.error(f"Ошибка получения филиалов: {e}")
        return []


# ─── HTTP-клиент к backend ───────────────────────────────────────────────────

async def api_get(path: str) -> dict | None:
    """GET-запрос к backend API."""
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.get(f"{BACKEND_URL}{path}")
            r.raise_for_status()
            return r.json()
    except Exception as e:
        logger.error(f"API GET {path}: {e}")
        return None


async def api_post(path: str, data: dict = None) -> dict | None:
    """POST-запрос к backend API."""
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(f"{BACKEND_URL}{path}", json=data or {})
            r.raise_for_status()
            return r.json()
    except Exception as e:
        logger.error(f"API POST {path}: {e}")
        return None


# ─── Вспомогательные функции ─────────────────────────────────────────────────

def is_authorized(user_id: int) -> bool:
    return user_id in authorized_users


def progress_bar(current: int, goal: int, width: int = 10) -> str:
    """Текстовый прогресс-бар: ▓▓▓▓░░░░░░ 40%"""
    if goal <= 0:
        return "░" * width + " 0%"
    pct = min(current / goal, 1.0)
    filled = round(pct * width)
    bar = "▓" * filled + "░" * (width - filled)
    return f"{bar} {round(pct * 100)}%"


def escape_md(text: str) -> str:
    """Экранирование спецсимволов для MarkdownV2."""
    special = r"_*[]()~`>#+-=|{}.!\\"
    result = ""
    for ch in str(text):
        if ch in special:
            result += "\\" + ch
        else:
            result += ch
    return result


SECTION_NAMES = {
    "morning": ("Утренние мероприятия", "/morning-events"),
    "field": ("Полевые выходы", "/field-visits"),
    "oneonone": ("One-on-One", "/one-on-one"),
    "weekly": ("Еженедельные показатели", "/weekly-metrics"),
    "masters": ("Планы мастеров", "/master-plans"),
    "reviews": ("Отзывы", "/reviews"),
    "newbie": ("Адаптация новичков", "/newbie-adaptation"),
    "summary": ("Итоговые отчёты", "/branch-summary"),
}


def build_xlsx_from_records(records: list[dict], sheet_name: str) -> bytes:
    """Создаёт Excel-файл из записей с форматированием."""
    if not records:
        return b""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    headers = list(records[0].keys())

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border

    data_align = Alignment(vertical="center", wrap_text=True)
    for ri, rec in enumerate(records, 2):
        for ci, h in enumerate(headers, 1):
            val = rec.get(h, "")
            if isinstance(val, str):
                try:
                    val = float(val) if "." in val else int(val)
                except (ValueError, TypeError):
                    pass
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment, c.border = data_align, border

    for ci, h in enumerate(headers, 1):
        max_len = len(str(h))
        for ri in range(2, len(records) + 2):
            v = ws.cell(row=ri, column=ci).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 3, 50)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Обработчики команд ─────────────────────────────────────────────────────

async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Приветствие и запрос пароля."""
    user_id = update.effective_user.id
    if is_authorized(user_id):
        await show_main_menu(update, ctx)
        return

    await update.message.reply_text(
        "🔐 *BarberCRM — Бот руководителя*\n\n"
        "Для доступа введите пароль:",
        parse_mode=ParseMode.MARKDOWN,
    )


async def handle_password(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Проверка пароля."""
    user_id = update.effective_user.id
    if is_authorized(user_id):
        # Уже авторизован — игнорируем текстовые сообщения
        return

    text = (update.message.text or "").strip()
    if text == BOT_PASSWORD:
        authorized_users.add(user_id)
        logger.info(f"✅ Авторизован: {user_id} ({update.effective_user.full_name})")
        await update.message.reply_text("✅ Доступ разрешён!")
        await show_main_menu(update, ctx)
    else:
        await update.message.reply_text("❌ Неверный пароль. Попробуйте ещё раз.")


# ─── Главное меню ────────────────────────────────────────────────────────────

async def show_main_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE, edit: bool = False):
    """Главное меню бота."""
    keyboard = [
        [InlineKeyboardButton("📊 Обзор всех филиалов", callback_data="overview")],
        [InlineKeyboardButton("🏢 Выбрать филиал", callback_data="select_branch")],
        [InlineKeyboardButton("📥 Выгрузить Excel", callback_data="export_select")],
        [InlineKeyboardButton("ℹ️ Информация", callback_data="info")],
    ]
    text = (
        "🏠 *Главное меню*\n\n"
        "Выберите действие:"
    )
    markup = InlineKeyboardMarkup(keyboard)

    if edit and update.callback_query:
        await update.callback_query.edit_message_text(
            text, reply_markup=markup, parse_mode=ParseMode.MARKDOWN
        )
    elif update.message:
        await update.message.reply_text(
            text, reply_markup=markup, parse_mode=ParseMode.MARKDOWN
        )
    elif update.callback_query:
        await update.callback_query.edit_message_text(
            text, reply_markup=markup, parse_mode=ParseMode.MARKDOWN
        )


# ─── Обзор всех филиалов ────────────────────────────────────────────────────

async def handle_overview(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Сводная таблица по всем филиалам."""
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("⏳ Загрузка данных по всем филиалам...")

    branches = get_branches_list()
    if not branches:
        await query.edit_message_text(
            "❌ Не удалось получить список филиалов.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 Повторить", callback_data="overview")],
                [InlineKeyboardButton("🏠 Меню", callback_data="main_menu")],
            ]),
        )
        return

    lines = ["📊 *Обзор филиалов*\n"]
    now = datetime.now()
    months_ru = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
    ]
    current_month = f"{months_ru[now.month - 1]} {now.year}"
    lines.append(f"📅 _{escape_md(current_month)}_\n")

    for br in branches:
        name = br["name"]
        data = await api_get(f"/dashboard-summary/{name}")
        if not data or not data.get("success"):
            lines.append(f"🏢 *{escape_md(name)}* — ⚠️ нет данных\n")
            continue

        summary = data["summary"]
        total_current = sum(v["current"] for v in summary.values())
        total_goal = sum(v["goal"] for v in summary.values())
        pct = round(total_current / total_goal * 100) if total_goal else 0

        emoji = "🟢" if pct >= 75 else "🟡" if pct >= 40 else "🔴"
        bar = progress_bar(total_current, total_goal, 8)

        lines.append(f"{emoji} *{escape_md(name)}*")
        lines.append(f"   👤 {escape_md(br.get('manager', '—'))}")
        lines.append(f"   {escape_md(bar)}  \\({total_current}/{total_goal}\\)\n")

    text = "\n".join(lines)

    await query.edit_message_text(
        text,
        parse_mode=ParseMode.MARKDOWN_V2,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 Обновить", callback_data="overview")],
            [InlineKeyboardButton("🏠 Меню", callback_data="main_menu")],
        ]),
    )


# ─── Выбор филиала ──────────────────────────────────────────────────────────

async def handle_select_branch(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Список филиалов для выбора."""
    query = update.callback_query
    await query.answer()

    branches = get_branches_list()
    if not branches:
        await query.edit_message_text(
            "❌ Не удалось загрузить список филиалов.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🏠 Меню", callback_data="main_menu")],
            ]),
        )
        return

    keyboard = []
    for br in branches:
        keyboard.append([
            InlineKeyboardButton(
                f"🏢 {br['name']}",
                callback_data=f"branch:{br['name']}",
            )
        ])
    keyboard.append([InlineKeyboardButton("🏠 Меню", callback_data="main_menu")])

    await query.edit_message_text(
        "🏢 *Выберите филиал:*",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


# ─── Дашборд филиала ────────────────────────────────────────────────────────

async def handle_branch_dashboard(update: Update, ctx: ContextTypes.DEFAULT_TYPE, branch_name: str):
    """Детальный дашборд одного филиала."""
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(f"⏳ Загрузка данных: {branch_name}...")

    data = await api_get(f"/dashboard-summary/{branch_name}")
    if not data or not data.get("success"):
        await query.edit_message_text(
            f"❌ Не удалось загрузить данные для *{branch_name}*",
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 Повторить", callback_data=f"branch:{branch_name}")],
                [InlineKeyboardButton("🏠 Меню", callback_data="main_menu")],
            ]),
        )
        return

    summary = data["summary"]
    now = datetime.now()
    months_ru = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
    ]
    current_month = f"{months_ru[now.month - 1]} {now.year}"

    lines = [
        f"🏢 *{escape_md(branch_name)}*",
        f"📅 _{escape_md(current_month)}_\n",
    ]

    icons = {
        "morning_events": "☀️",
        "field_visits": "🚶",
        "one_on_one": "🤝",
        "master_plans": "📋",
        "weekly_reports": "📈",
        "reviews": "⭐",
        "new_employees": "👶",
    }

    for key, vals in summary.items():
        icon = icons.get(key, "📌")
        label = vals.get("label", key)
        cur = vals["current"]
        goal = vals["goal"]
        bar = progress_bar(cur, goal, 8)
        lines.append(f"{icon} *{escape_md(label)}*")
        lines.append(f"   {escape_md(bar)}  {cur}/{goal}\n")

    total_cur = sum(v["current"] for v in summary.values())
    total_goal = sum(v["goal"] for v in summary.values())
    total_pct = round(total_cur / total_goal * 100) if total_goal else 0
    emoji = "🟢" if total_pct >= 75 else "🟡" if total_pct >= 40 else "🔴"
    lines.append(f"{emoji} *Общий прогресс: {total_pct}%*")

    text = "\n".join(lines)

    # Кнопки разделов
    keyboard = [
        [
            InlineKeyboardButton("☀️ Утренние", callback_data=f"sec:morning:{branch_name}"),
            InlineKeyboardButton("🚶 Полевые", callback_data=f"sec:field:{branch_name}"),
        ],
        [
            InlineKeyboardButton("🤝 One-on-One", callback_data=f"sec:oneonone:{branch_name}"),
            InlineKeyboardButton("📈 Еженед.", callback_data=f"sec:weekly:{branch_name}"),
        ],
        [
            InlineKeyboardButton("📋 Планы", callback_data=f"sec:masters:{branch_name}"),
            InlineKeyboardButton("⭐ Отзывы", callback_data=f"sec:reviews:{branch_name}"),
        ],
        [
            InlineKeyboardButton("👶 Новички", callback_data=f"sec:newbie:{branch_name}"),
            InlineKeyboardButton("📊 Итоги", callback_data=f"sec:summary:{branch_name}"),
        ],
        [InlineKeyboardButton("📥 Выгрузить всё в Excel", callback_data=f"export_all:{branch_name}")],
        [
            InlineKeyboardButton("🔄 Обновить", callback_data=f"branch:{branch_name}"),
            InlineKeyboardButton("🏠 Меню", callback_data="main_menu"),
        ],
    ]

    await query.edit_message_text(
        text,
        parse_mode=ParseMode.MARKDOWN_V2,
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


# ─── Просмотр раздела ───────────────────────────────────────────────────────

async def handle_section(update: Update, ctx: ContextTypes.DEFAULT_TYPE, section: str, branch_name: str):
    """Показать последние записи раздела."""
    query = update.callback_query
    await query.answer()

    sec_info = SECTION_NAMES.get(section)
    if not sec_info:
        await query.edit_message_text("❌ Неизвестный раздел")
        return

    sec_label, api_path = sec_info
    await query.edit_message_text(f"⏳ Загрузка: {sec_label}...")

    data = await api_get(f"{api_path}/{branch_name}")
    if not data or not data.get("success"):
        await query.edit_message_text(
            f"❌ Не удалось загрузить *{sec_label}*",
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("◀️ Назад", callback_data=f"branch:{branch_name}")],
            ]),
        )
        return

    records = data.get("data", [])
    if not records:
        await query.edit_message_text(
            f"📭 *{escape_md(sec_label)}*\n\nНет записей для *{escape_md(branch_name)}*",
            parse_mode=ParseMode.MARKDOWN_V2,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("◀️ Назад", callback_data=f"branch:{branch_name}")],
                [InlineKeyboardButton("🏠 Меню", callback_data="main_menu")],
            ]),
        )
        return

    # Форматируем последние 5 записей
    lines = [f"📋 *{escape_md(sec_label)}*  \\|  *{escape_md(branch_name)}*"]
    lines.append(f"Всего записей: {len(records)}\n")

    last_records = records[:5]

    if section == "morning":
        for r in last_records:
            date = r.get("Дата", r.get("Дата отправки", "—"))
            etype = r.get("Тип мероприятия", "—")
            parts = r.get("Участники", "—")
            eff = r.get("Эффективность", "—")
            lines.append(f"📅 *{escape_md(str(date))}*")
            lines.append(f"   {escape_md(str(etype))} • 👥 {parts} • ⚡ {eff}/5")
            comment = r.get("Комментарий", "")
            if comment:
                lines.append(f"   💬 _{escape_md(str(comment)[:80])}_")
            lines.append("")

    elif section == "field":
        for r in last_records:
            date = r.get("Дата", "—")
            master = r.get("Имя мастера", "—")
            avg = r.get("Общая оценка", "—")
            lines.append(f"📅 *{escape_md(str(date))}* — {escape_md(str(master))}")
            lines.append(f"   Общая оценка: *{avg}/10*")
            errors = r.get("Ошибки", "")
            if errors:
                lines.append(f"   ⚠️ _{escape_md(str(errors)[:80])}_")
            lines.append("")

    elif section == "oneonone":
        for r in last_records:
            date = r.get("Дата", "—")
            master = r.get("Имя мастера", "—")
            goal = r.get("Цель", "—")
            lines.append(f"📅 *{escape_md(str(date))}* — {escape_md(str(master))}")
            lines.append(f"   🎯 {escape_md(str(goal)[:100])}")
            results = r.get("Результаты", "")
            if results:
                lines.append(f"   ✅ _{escape_md(str(results)[:80])}_")
            lines.append("")

    elif section == "weekly":
        for r in last_records:
            period = r.get("Период", "—")
            chk_p = r.get("Средний чек (план)", "—")
            chk_f = r.get("Средний чек (факт)", "—")
            cos_p = r.get("Косметика (план)", "—")
            cos_f = r.get("Косметика (факт)", "—")
            lines.append(f"📅 *{escape_md(str(period))}*")
            lines.append(f"   💰 Ср\\.чек: {chk_f}/{chk_p}")
            lines.append(f"   🧴 Косметика: {cos_f}/{cos_p}")
            lines.append("")

    elif section == "masters":
        for r in last_records:
            month = r.get("Месяц", "—")
            master = r.get("Имя мастера", "—")
            chk_p = r.get("Средний чек (план)", 0)
            chk_f = r.get("Средний чек (факт)", 0)
            sales_p = r.get("Продажи (план)", 0)
            sales_f = r.get("Продажи (факт)", 0)
            lines.append(f"👤 *{escape_md(str(master))}* \\| {escape_md(str(month))}")
            lines.append(f"   💰 Ср\\.чек: {chk_f}/{chk_p}")
            lines.append(f"   🛒 Продажи: {sales_f}/{sales_p}")
            lines.append("")

    elif section == "reviews":
        for r in last_records:
            week = r.get("Неделя", "—")
            plan_v = r.get("План", "—")
            fact_v = r.get("Факт", "—")
            lines.append(f"📅 Неделя *{escape_md(str(week))}*: {fact_v}/{plan_v}")

    elif section == "newbie":
        for r in last_records:
            name = r.get("Имя", "—")
            status = r.get("Статус", "—")
            start = r.get("Дата начала", "—")
            lines.append(f"👶 *{escape_md(str(name))}* — {escape_md(str(status))}")
            lines.append(f"   📅 Начало: {escape_md(str(start))}")
            lines.append("")

    elif section == "summary":
        for r in last_records:
            metric = r.get("Метрика", "—")
            cur = r.get("Текущее количество", "—")
            goal = r.get("Цель на месяц", "—")
            pct = r.get("Выполнение %", "—")
            month = r.get("Месяц", "")
            lines.append(f"• *{escape_md(str(metric))}*: {cur}/{goal} \\({pct}%\\)")

    else:
        # Универсальный вывод
        for r in last_records:
            items = list(r.items())[:4]
            line_parts = [f"{escape_md(str(v))}" for k, v in items]
            lines.append(" • ".join(line_parts))

    if len(records) > 5:
        lines.append(f"\n_\\.\\.\\.и ещё {len(records) - 5} записей_")

    text = "\n".join(lines)

    # Обрезаем если слишком длинный
    if len(text) > 4000:
        text = text[:3950] + "\n\n_\\.\\.\\. сообщение обрезано_"

    keyboard = [
        [InlineKeyboardButton(
            f"📥 Выгрузить {sec_label} в Excel",
            callback_data=f"export_sec:{section}:{branch_name}",
        )],
        [InlineKeyboardButton("◀️ Назад к филиалу", callback_data=f"branch:{branch_name}")],
        [InlineKeyboardButton("🏠 Меню", callback_data="main_menu")],
    ]

    await query.edit_message_text(
        text,
        parse_mode=ParseMode.MARKDOWN_V2,
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


# ─── Выгрузка Excel ─────────────────────────────────────────────────────────

async def handle_export_select(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Выбор филиала для экспорта."""
    query = update.callback_query
    await query.answer()

    branches = get_branches_list()
    if not branches:
        await query.edit_message_text("❌ Нет филиалов.")
        return

    keyboard = []
    for br in branches:
        keyboard.append([
            InlineKeyboardButton(
                f"📥 {br['name']}",
                callback_data=f"export_all:{br['name']}",
            )
        ])
    keyboard.append([InlineKeyboardButton("🏠 Меню", callback_data="main_menu")])

    await query.edit_message_text(
        "📥 *Выгрузка в Excel*\n\nВыберите филиал:",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


async def handle_export_all(update: Update, ctx: ContextTypes.DEFAULT_TYPE, branch_name: str):
    """Выгрузка ВСЕХ разделов филиала одним Excel-файлом (multi-sheet)."""
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(f"⏳ Формирование Excel для {branch_name}...")

    wb = Workbook()
    wb.remove(wb.active)  # Удаляем дефолтный лист

    has_data = False
    for sec_key, (sec_label, api_path) in SECTION_NAMES.items():
        data = await api_get(f"{api_path}/{branch_name}")
        if data and data.get("success") and data.get("data"):
            records = data["data"]
            ws = wb.create_sheet(title=sec_label[:31])

            headers = list(records[0].keys())
            hdr_font = Font(bold=True, color="FFFFFF", size=11)
            hdr_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
            hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border

            data_align = Alignment(vertical="center", wrap_text=True)
            for ri, rec in enumerate(records, 2):
                for ci, h in enumerate(headers, 1):
                    val = rec.get(h, "")
                    if isinstance(val, str):
                        try:
                            val = float(val) if "." in val else int(val)
                        except (ValueError, TypeError):
                            pass
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.alignment, c.border = data_align, border

            for ci, h in enumerate(headers, 1):
                max_len = len(str(h))
                for ri in range(2, min(len(records) + 2, 100)):
                    v = ws.cell(row=ri, column=ci).value
                    if v:
                        max_len = max(max_len, len(str(v)))
                ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 3, 50)

            ws.freeze_panes = "A2"
            has_data = True

    if not has_data:
        await query.edit_message_text(
            f"📭 Нет данных для выгрузки по филиалу *{branch_name}*",
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🏠 Меню", callback_data="main_menu")],
            ]),
        )
        return

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = branch_name.replace(" ", "_")
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"BarberCRM_{safe_name}_{date_str}.xlsx"

    await query.message.reply_document(
        document=buf,
        filename=filename,
        caption=f"📥 Данные филиала: {branch_name}\n📅 {date_str}",
    )

    await query.message.reply_text(
        "✅ Файл отправлен!",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("◀️ Назад к филиалу", callback_data=f"branch:{branch_name}")],
            [InlineKeyboardButton("🏠 Меню", callback_data="main_menu")],
        ]),
    )


async def handle_export_section(update: Update, ctx: ContextTypes.DEFAULT_TYPE, section: str, branch_name: str):
    """Выгрузка одного раздела в Excel."""
    query = update.callback_query
    await query.answer()

    sec_info = SECTION_NAMES.get(section)
    if not sec_info:
        await query.edit_message_text("❌ Неизвестный раздел")
        return

    sec_label, api_path = sec_info
    await query.edit_message_text(f"⏳ Выгрузка: {sec_label}...")

    data = await api_get(f"{api_path}/{branch_name}")
    if not data or not data.get("success") or not data.get("data"):
        await query.edit_message_text(
            f"📭 Нет данных в разделе *{sec_label}*",
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("◀️ Назад", callback_data=f"branch:{branch_name}")],
            ]),
        )
        return

    records = data["data"]
    xlsx_bytes = build_xlsx_from_records(records, sec_label)

    buf = io.BytesIO(xlsx_bytes)
    safe_name = branch_name.replace(" ", "_")
    safe_sec = sec_label.replace(" ", "_")
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"{safe_sec}_{safe_name}_{date_str}.xlsx"

    await query.message.reply_document(
        document=buf,
        filename=filename,
        caption=f"📥 {sec_label} — {branch_name}\n📅 {date_str}\n📊 {len(records)} записей",
    )

    await query.message.reply_text(
        "✅ Файл отправлен!",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("◀️ Назад к разделу", callback_data=f"sec:{section}:{branch_name}")],
            [InlineKeyboardButton("◀️ К филиалу", callback_data=f"branch:{branch_name}")],
            [InlineKeyboardButton("🏠 Меню", callback_data="main_menu")],
        ]),
    )


# ─── Информация ─────────────────────────────────────────────────────────────

async def handle_info(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Информация о боте и системе."""
    query = update.callback_query
    await query.answer()

    health = await api_get("/health")
    status = "✅ Работает" if health else "❌ Недоступен"
    version = health.get("version", "—") if health else "—"

    branches = get_branches_list()

    text = (
        "ℹ️ *BarberCRM — Бот руководителя*\n\n"
        f"🔗 Backend: {status}\n"
        f"📌 Версия API: {version}\n"
        f"🏢 Филиалов: {len(branches)}\n\n"
        "📊 *Возможности:*\n"
        "• Обзор всех филиалов\n"
        "• Детальный дашборд каждого филиала\n"
        "• Просмотр всех разделов данных\n"
        "• Выгрузка в Excel (по разделу или всё сразу)\n\n"
        "🔄 Данные обновляются в реальном времени из Google Sheets"
    )

    await query.edit_message_text(
        text,
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🏠 Меню", callback_data="main_menu")],
        ]),
    )


# ─── Главный маршрутизатор callback ─────────────────────────────────────────

async def callback_router(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Маршрутизация всех callback_query."""
    query = update.callback_query
    user_id = query.from_user.id

    if not is_authorized(user_id):
        await query.answer("🔐 Необходима авторизация. Отправьте /start", show_alert=True)
        return

    data = query.data

    if data == "main_menu":
        await show_main_menu(update, ctx, edit=True)

    elif data == "overview":
        await handle_overview(update, ctx)

    elif data == "select_branch":
        await handle_select_branch(update, ctx)

    elif data.startswith("branch:"):
        branch_name = data.split(":", 1)[1]
        await handle_branch_dashboard(update, ctx, branch_name)

    elif data.startswith("sec:"):
        parts = data.split(":", 2)
        if len(parts) == 3:
            _, section, branch_name = parts
            await handle_section(update, ctx, section, branch_name)

    elif data == "export_select":
        await handle_export_select(update, ctx)

    elif data.startswith("export_all:"):
        branch_name = data.split(":", 1)[1]
        await handle_export_all(update, ctx, branch_name)

    elif data.startswith("export_sec:"):
        parts = data.split(":", 2)
        if len(parts) == 3:
            _, section, branch_name = parts
            await handle_export_section(update, ctx, section, branch_name)

    elif data == "info":
        await handle_info(update, ctx)

    else:
        await query.answer("❓ Неизвестная команда")


# ─── Запуск ──────────────────────────────────────────────────────────────────

def main():
    if not BOT_TOKEN:
        logger.error("❌ TELEGRAM_BOT_TOKEN не задан!")
        return

    logger.info("🤖 Запуск BarberCRM Bot...")
    logger.info(f"🔗 Backend URL: {BACKEND_URL}")

    app = Application.builder().token(BOT_TOKEN).build()

    # Команды
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("menu", lambda u, c: show_main_menu(u, c)))

    # Callback кнопки
    app.add_handler(CallbackQueryHandler(callback_router))

    # Текстовые сообщения (для пароля)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_password))

    logger.info("✅ Бот запущен и ожидает сообщения...")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
