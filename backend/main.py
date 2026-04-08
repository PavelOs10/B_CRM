from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import json, os, hashlib, secrets, logging, time, smtplib, io, sqlite3
from datetime import datetime, timedelta
from contextlib import contextmanager
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="BarberCRM API", version="5.1.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# ============= CONFIG =============
REPORT_EMAIL_TO = os.getenv('REPORT_EMAIL_TO', '')
SMTP_HOST = os.getenv('SMTP_HOST', '')
SMTP_PORT = int(os.getenv('SMTP_PORT', '587'))
SMTP_USER = os.getenv('SMTP_USER', '')
SMTP_PASSWORD = os.getenv('SMTP_PASSWORD', '')
SMTP_USE_SSL = os.getenv('SMTP_USE_SSL', 'false').lower() == 'true'
ADMIN_USERNAME = os.getenv('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD_HASH = hashlib.sha256(os.getenv('ADMIN_PASSWORD', 'admin').encode()).hexdigest()
DB_PATH = os.getenv('DB_PATH', '/app/data/barbercrm.db')

BRANCH_GOALS = {"morning_events": 16, "field_visits": 4, "one_on_one": 6, "weekly_reports": 4, "master_plans": 10, "reviews": 52, "new_employees": 10}

# ============= DATABASE =============
@contextmanager
def get_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def init_db():
    with get_db() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS branches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL, address TEXT NOT NULL,
            manager_name TEXT NOT NULL, manager_phone TEXT NOT NULL,
            password_hash TEXT NOT NULL, token TEXT NOT NULL, created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS morning_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT, branch_name TEXT NOT NULL,
            submitted_at TEXT NOT NULL, date TEXT NOT NULL, week INTEGER NOT NULL,
            event_type TEXT NOT NULL, participants INTEGER NOT NULL,
            efficiency INTEGER NOT NULL, comment TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS field_visits (
            id INTEGER PRIMARY KEY AUTOINCREMENT, branch_name TEXT NOT NULL,
            submitted_at TEXT NOT NULL, date TEXT NOT NULL, master_name TEXT NOT NULL,
            haircut_quality INTEGER NOT NULL, service_quality INTEGER NOT NULL,
            additional_services_comment TEXT DEFAULT '', additional_services_rating INTEGER NOT NULL,
            cosmetics_comment TEXT DEFAULT '', cosmetics_rating INTEGER NOT NULL,
            standards_comment TEXT DEFAULT '', standards_rating INTEGER NOT NULL,
            errors_comment TEXT DEFAULT '', next_check_date TEXT DEFAULT '', average_rating REAL NOT NULL
        );
        CREATE TABLE IF NOT EXISTS one_on_one (
            id INTEGER PRIMARY KEY AUTOINCREMENT, branch_name TEXT NOT NULL,
            submitted_at TEXT NOT NULL, date TEXT NOT NULL, master_name TEXT NOT NULL,
            goal TEXT NOT NULL, results TEXT NOT NULL, development_plan TEXT NOT NULL,
            indicator TEXT NOT NULL, next_meeting_date TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS weekly_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT, branch_name TEXT NOT NULL,
            submitted_at TEXT NOT NULL, period TEXT NOT NULL,
            average_check_plan REAL NOT NULL, average_check_fact REAL NOT NULL,
            cosmetics_plan REAL NOT NULL, cosmetics_fact REAL NOT NULL,
            additional_services_plan REAL NOT NULL, additional_services_fact REAL NOT NULL
        );
        CREATE TABLE IF NOT EXISTS master_plans (
            id INTEGER PRIMARY KEY AUTOINCREMENT, branch_name TEXT NOT NULL,
            submitted_at TEXT NOT NULL, month TEXT NOT NULL, master_name TEXT NOT NULL,
            average_check_plan REAL NOT NULL, average_check_fact REAL NOT NULL,
            additional_services_plan INTEGER NOT NULL, additional_services_fact INTEGER NOT NULL,
            sales_plan REAL NOT NULL, sales_fact REAL NOT NULL,
            salary_plan REAL NOT NULL, salary_fact REAL NOT NULL
        );
        CREATE TABLE IF NOT EXISTS reviews (
            id INTEGER PRIMARY KEY AUTOINCREMENT, branch_name TEXT NOT NULL,
            submitted_at TEXT NOT NULL, week TEXT NOT NULL, manager_name TEXT NOT NULL,
            plan INTEGER NOT NULL DEFAULT 13, fact INTEGER NOT NULL, monthly_target INTEGER NOT NULL DEFAULT 52
        );
        CREATE TABLE IF NOT EXISTS newbie_adaptation (
            id INTEGER PRIMARY KEY AUTOINCREMENT, branch_name TEXT NOT NULL,
            submitted_at TEXT NOT NULL, start_date TEXT NOT NULL, name TEXT NOT NULL,
            haircut_practice TEXT NOT NULL, service_standards TEXT NOT NULL,
            hygiene_sanitation TEXT NOT NULL, additional_services TEXT NOT NULL,
            cosmetics_sales TEXT NOT NULL, iclient_basics TEXT NOT NULL, status TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS branch_summaries (
            id INTEGER PRIMARY KEY AUTOINCREMENT, branch_name TEXT NOT NULL,
            submitted_at TEXT NOT NULL, manager TEXT NOT NULL, month TEXT NOT NULL,
            metric TEXT NOT NULL, current_value INTEGER NOT NULL,
            goal_value INTEGER NOT NULL, percentage REAL NOT NULL
        );
        """)
    logger.info("✅ БД инициализирована")

# ============= UTILS =============
def hash_password(p): return hashlib.sha256(p.encode()).hexdigest()
def generate_token(): return secrets.token_urlsafe(32)

def parse_date_flexible(date_str):
    s = str(date_str).strip()
    if not s: return None
    for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"]:
        try: return datetime.strptime(s.split()[0], fmt)
        except ValueError: continue
    return None

def get_month_ru(dt):
    m = ['Январь','Февраль','Март','Апрель','Май','Июнь','Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']
    return f"{m[dt.month-1]} {dt.year}"

def current_month_ru(): return get_month_ru(datetime.now())

def count_for_month(rows, field, month):
    c = 0
    for r in rows:
        dt = parse_date_flexible(str(r[field]))
        if dt and get_month_ru(dt) == month: c += 1
    return c

def sum_reviews_month(rows, month):
    t = 0
    for r in rows:
        dt = parse_date_flexible(str(r['submitted_at']))
        if dt and get_month_ru(dt) == month: t += int(r['fact'] or 0)
    return t

def get_period_dates(period_type, custom_date=None):
    """Возвращает (start, end, label) для фильтрации"""
    now = datetime.now()
    if period_type == "today":
        s = now.replace(hour=0, minute=0, second=0, microsecond=0)
        e = now.replace(hour=23, minute=59, second=59)
        return s, e, f"Сегодня ({now.strftime('%d.%m.%Y')})"
    elif period_type == "yesterday":
        y = now - timedelta(days=1)
        s = y.replace(hour=0, minute=0, second=0, microsecond=0)
        e = y.replace(hour=23, minute=59, second=59)
        return s, e, f"Вчера ({y.strftime('%d.%m.%Y')})"
    elif period_type == "week":
        wd = now.weekday()
        s = (now - timedelta(days=wd)).replace(hour=0, minute=0, second=0, microsecond=0)
        e = (s + timedelta(days=6)).replace(hour=23, minute=59, second=59)
        return s, e, f"Неделя ({s.strftime('%d.%m')}–{e.strftime('%d.%m.%Y')})"
    elif period_type == "month":
        s = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        e = (s + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
        return s, e, current_month_ru()
    elif period_type == "quarter":
        q = (now.month - 1) // 3
        s = datetime(now.year, q * 3 + 1, 1)
        e_month = q * 3 + 3
        e = datetime(now.year, e_month, 1) + timedelta(days=32)
        e = e.replace(day=1) - timedelta(seconds=1)
        return s, e, f"Q{q+1} {now.year}"
    elif period_type == "year":
        s = datetime(now.year, 1, 1)
        e = datetime(now.year, 12, 31, 23, 59, 59)
        return s, e, f"{now.year} год"
    elif period_type == "day" and custom_date:
        t = datetime.strptime(custom_date, "%Y-%m-%d")
        return t.replace(hour=0,minute=0,second=0), t.replace(hour=23,minute=59,second=59), t.strftime("%d.%m.%Y")
    else:
        return datetime(2020,1,1), datetime(2099,12,31), "Весь период"

def filter_by_period(rows, start, end):
    result = []
    for r in rows:
        d = dict(r)
        ds = d.get('Дата отправки','') or d.get('submitted_at','')
        dt = parse_date_flexible(str(ds))
        if dt and start.date() <= dt.date() <= end.date():
            result.append(d)
    return result

# ============= MODELS =============
class BranchRegister(BaseModel):
    name: str; address: str; manager_name: str; manager_phone: str; password: str

class BranchUpdate(BaseModel):
    address: Optional[str] = None; manager_name: Optional[str] = None
    manager_phone: Optional[str] = None; password: Optional[str] = None

class LoginRequest(BaseModel):
    name: str; password: str

class AdminLoginRequest(BaseModel):
    username: str; password: str

class MorningEvent(BaseModel):
    week: int = Field(..., ge=1, le=53); date: str; event_type: str
    participants: int = Field(..., ge=0, le=100); efficiency: int = Field(..., ge=1, le=5)
    comment: Optional[str] = ""

class FieldVisit(BaseModel):
    date: str; master_name: str
    haircut_quality: int = Field(..., ge=1, le=10); service_quality: int = Field(..., ge=1, le=10)
    additional_services_comment: str = ""; additional_services_rating: int = Field(..., ge=1, le=10)
    cosmetics_comment: str = ""; cosmetics_rating: int = Field(..., ge=1, le=10)
    standards_comment: str = ""; standards_rating: int = Field(..., ge=1, le=10)
    errors_comment: str = ""; next_check_date: Optional[str] = ""

class OneOnOneMeeting(BaseModel):
    date: str; master_name: str; goal: str; results: str
    development_plan: str; indicator: str; next_meeting_date: Optional[str] = ""

class WeeklyMetrics(BaseModel):
    period: str
    average_check_plan: float = Field(..., ge=0); average_check_fact: float = Field(..., ge=0)
    cosmetics_plan: float = Field(..., ge=0); cosmetics_fact: float = Field(..., ge=0)
    additional_services_plan: float = Field(..., ge=0); additional_services_fact: float = Field(..., ge=0)

class NewbieAdaptation(BaseModel):
    start_date: str; name: str; haircut_practice: str; service_standards: str
    hygiene_sanitation: str; additional_services: str; cosmetics_sales: str
    iclient_basics: str; status: str

class MasterPlan(BaseModel):
    month: str; master_name: str
    average_check_plan: float = Field(..., ge=0); average_check_fact: float = Field(..., ge=0)
    additional_services_plan: int = Field(..., ge=0); additional_services_fact: int = Field(..., ge=0)
    sales_plan: float = Field(..., ge=0); sales_fact: float = Field(..., ge=0)
    salary_plan: float = Field(..., ge=0); salary_fact: float = Field(..., ge=0)

class Reviews(BaseModel):
    week: str; manager_name: str; plan: int = 13; fact: int = Field(..., ge=0); monthly_target: int = 52

class BranchSummary(BaseModel):
    manager: str; month: str

class EmailReportRequest(BaseModel):
    period_type: str; custom_date: Optional[str] = None

# ============= STARTUP =============
@app.on_event("startup")
def startup(): init_db()

@app.get("/health")
def health(): return {"status": "healthy", "version": "5.1.0"}

# ============= AUTH =============
@app.post("/register")
def register_branch(b: BranchRegister):
    with get_db() as conn:
        if conn.execute("SELECT id FROM branches WHERE name=?", (b.name,)).fetchone():
            raise HTTPException(400, "Филиал с таким названием уже существует")
        token = generate_token()
        conn.execute("INSERT INTO branches (name,address,manager_name,manager_phone,password_hash,token,created_at) VALUES (?,?,?,?,?,?,?)",
            (b.name, b.address, b.manager_name, b.manager_phone, hash_password(b.password), token, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    return {"success": True, "message": "Филиал зарегистрирован", "token": token, "branch_name": b.name}

@app.post("/login")
def login(r: LoginRequest):
    with get_db() as conn:
        br = conn.execute("SELECT * FROM branches WHERE name=?", (r.name,)).fetchone()
        if not br: raise HTTPException(401, "Неверное название филиала")
        if br['password_hash'] != hash_password(r.password): raise HTTPException(401, "Неверный пароль")
        return {"success": True, "token": br['token'], "branch": {"name": r.name, "manager": br['manager_name']}}

@app.post("/admin/login")
def admin_login(r: AdminLoginRequest):
    if r.username != ADMIN_USERNAME or hash_password(r.password) != ADMIN_PASSWORD_HASH:
        raise HTTPException(401, "Неверный логин или пароль")
    return {"success": True, "token": generate_token(), "role": "admin"}

@app.get("/branches")
def get_branches():
    with get_db() as conn:
        rows = conn.execute("SELECT name FROM branches ORDER BY name").fetchall()
    return {"success": True, "branches": [r['name'] for r in rows]}

@app.get("/branches/details")
def get_branches_details():
    with get_db() as conn:
        rows = conn.execute("SELECT name, address, manager_name, manager_phone, created_at FROM branches ORDER BY name").fetchall()
    return {"success": True, "branches": [dict(r) for r in rows]}

# ============= ADMIN: УПРАВЛЕНИЕ ФИЛИАЛАМИ =============
@app.put("/admin/branches/{branch_name}")
def admin_update_branch(branch_name: str, data: BranchUpdate):
    with get_db() as conn:
        br = conn.execute("SELECT id FROM branches WHERE name=?", (branch_name,)).fetchone()
        if not br: raise HTTPException(404, "Филиал не найден")
        if data.manager_name: conn.execute("UPDATE branches SET manager_name=? WHERE name=?", (data.manager_name, branch_name))
        if data.manager_phone: conn.execute("UPDATE branches SET manager_phone=? WHERE name=?", (data.manager_phone, branch_name))
        if data.address: conn.execute("UPDATE branches SET address=? WHERE name=?", (data.address, branch_name))
        if data.password: conn.execute("UPDATE branches SET password_hash=? WHERE name=?", (hash_password(data.password), branch_name))
    return {"success": True, "message": f"Филиал '{branch_name}' обновлён"}

@app.delete("/admin/branches/{branch_name}")
def admin_delete_branch(branch_name: str):
    with get_db() as conn:
        br = conn.execute("SELECT id FROM branches WHERE name=?", (branch_name,)).fetchone()
        if not br: raise HTTPException(404, "Филиал не найден")
        for t in ["morning_events","field_visits","one_on_one","weekly_metrics","master_plans","reviews","newbie_adaptation","branch_summaries"]:
            conn.execute(f"DELETE FROM {t} WHERE branch_name=?", (branch_name,))
        conn.execute("DELETE FROM branches WHERE name=?", (branch_name,))
    return {"success": True, "message": f"Филиал '{branch_name}' и все его данные удалены"}

# ============= GENERIC CRUD HELPERS =============
SECTION_CONFIG = {
    "morning-events": {
        "table": "morning_events",
        "select": "id, submitted_at as 'Дата отправки', date as 'Дата', week as 'Неделя', event_type as 'Тип мероприятия', participants as 'Участники', efficiency as 'Эффективность', comment as 'Комментарий'",
        "update_fields": {"date":"date","week":"week","event_type":"event_type","participants":"participants","efficiency":"efficiency","comment":"comment"},
    },
    "field-visits": {
        "table": "field_visits",
        "select": "id, submitted_at as 'Дата отправки', date as 'Дата', master_name as 'Имя мастера', haircut_quality as 'Качество стрижки', service_quality as 'Качество обслуживания', additional_services_comment as 'Доп. услуги (комм.)', additional_services_rating as 'Доп. услуги (оценка)', cosmetics_comment as 'Косметика (комм.)', cosmetics_rating as 'Косметика (оценка)', standards_comment as 'Стандарты (комм.)', standards_rating as 'Стандарты (оценка)', errors_comment as 'Ошибки', next_check_date as 'Дата след. проверки', average_rating as 'Общая оценка'",
    },
    "one-on-one": {
        "table": "one_on_one",
        "select": "id, submitted_at as 'Дата отправки', date as 'Дата', master_name as 'Имя мастера', goal as 'Цель', results as 'Результаты', development_plan as 'План развития', indicator as 'Показатель', next_meeting_date as 'Дата след. встречи'",
    },
    "weekly-metrics": {
        "table": "weekly_metrics",
        "select": "id, submitted_at as 'Дата отправки', period as 'Период', average_check_plan as 'Средний чек (план)', average_check_fact as 'Средний чек (факт)', cosmetics_plan as 'Косметика (план)', cosmetics_fact as 'Косметика (факт)', additional_services_plan as 'Доп. услуги (план)', additional_services_fact as 'Доп. услуги (факт)'",
    },
    "master-plans": {
        "table": "master_plans",
        "select": "id, submitted_at as 'Дата отправки', month as 'Месяц', master_name as 'Имя мастера', average_check_plan as 'Средний чек (план)', average_check_fact as 'Средний чек (факт)', additional_services_plan as 'Доп. услуги (план)', additional_services_fact as 'Доп. услуги (факт)', sales_plan as 'Продажи (план)', sales_fact as 'Продажи (факт)', salary_plan as 'ЗП (план)', salary_fact as 'ЗП (факт)'",
    },
    "reviews": {
        "table": "reviews",
        "select": "id, submitted_at as 'Дата отправки', week as 'Неделя', manager_name as 'Имя руководителя', plan as 'План', fact as 'Факт', monthly_target as 'Месячная цель'",
    },
    "newbie-adaptation": {
        "table": "newbie_adaptation",
        "select": "id, submitted_at as 'Дата отправки', start_date as 'Дата начала', name as 'Имя', haircut_practice as 'Практика стрижки', service_standards as 'Стандарты обслуживания', hygiene_sanitation as 'Гигиена/санитария', additional_services as 'Доп. услуги', cosmetics_sales as 'Продажи косметики', iclient_basics as 'Основы iClient', status as 'Статус'",
    },
    "branch-summary": {
        "table": "branch_summaries",
        "select": "id, submitted_at as 'Дата отправки', manager as 'Руководитель', month as 'Месяц', metric as 'Метрика', current_value as 'Текущее количество', goal_value as 'Цель на месяц', percentage as 'Выполнение %'",
    },
}

def get_section_data(branch_name, section):
    cfg = SECTION_CONFIG.get(section)
    if not cfg: raise HTTPException(400, f"Неизвестная секция: {section}")
    with get_db() as conn:
        rows = conn.execute(f"SELECT {cfg['select']} FROM {cfg['table']} WHERE branch_name=? ORDER BY id DESC", (branch_name,)).fetchall()
    return {"success": True, "data": [dict(r) for r in rows]}

# ============= УНИВЕРСАЛЬНОЕ РЕДАКТИРОВАНИЕ И УДАЛЕНИЕ =============
@app.put("/record/{section}/{record_id}")
def update_record(section: str, record_id: int, data: Dict[str, Any]):
    """Универсальное обновление записи по id. Принимает JSON с полями для обновления."""
    cfg = SECTION_CONFIG.get(section)
    if not cfg: raise HTTPException(400, f"Неизвестная секция: {section}")
    table = cfg['table']
    
    # Маппинг русских названий колонок → реальные колонки в БД
    COLUMN_MAP = {
        "morning_events": {"Дата":"date","Неделя":"week","Тип мероприятия":"event_type","Участники":"participants","Эффективность":"efficiency","Комментарий":"comment"},
        "field_visits": {"Дата":"date","Имя мастера":"master_name","Качество стрижки":"haircut_quality","Качество обслуживания":"service_quality","Доп. услуги (комм.)":"additional_services_comment","Доп. услуги (оценка)":"additional_services_rating","Косметика (комм.)":"cosmetics_comment","Косметика (оценка)":"cosmetics_rating","Стандарты (комм.)":"standards_comment","Стандарты (оценка)":"standards_rating","Ошибки":"errors_comment","Дата след. проверки":"next_check_date"},
        "one_on_one": {"Дата":"date","Имя мастера":"master_name","Цель":"goal","Результаты":"results","План развития":"development_plan","Показатель":"indicator","Дата след. встречи":"next_meeting_date"},
        "weekly_metrics": {"Период":"period","Средний чек (план)":"average_check_plan","Средний чек (факт)":"average_check_fact","Косметика (план)":"cosmetics_plan","Косметика (факт)":"cosmetics_fact","Доп. услуги (план)":"additional_services_plan","Доп. услуги (факт)":"additional_services_fact"},
        "master_plans": {"Месяц":"month","Имя мастера":"master_name","Средний чек (план)":"average_check_plan","Средний чек (факт)":"average_check_fact","Доп. услуги (план)":"additional_services_plan","Доп. услуги (факт)":"additional_services_fact","Продажи (план)":"sales_plan","Продажи (факт)":"sales_fact","ЗП (план)":"salary_plan","ЗП (факт)":"salary_fact"},
        "reviews": {"Неделя":"week","Имя руководителя":"manager_name","План":"plan","Факт":"fact","Месячная цель":"monthly_target"},
        "newbie_adaptation": {"Дата начала":"start_date","Имя":"name","Практика стрижки":"haircut_practice","Стандарты обслуживания":"service_standards","Гигиена/санитария":"hygiene_sanitation","Доп. услуги":"additional_services","Продажи косметики":"cosmetics_sales","Основы iClient":"iclient_basics","Статус":"status"},
    }
    
    col_map = COLUMN_MAP.get(table, {})
    sets = []
    vals = []
    for key, val in data.items():
        if key in ('id', 'Дата отправки', 'branch_name', 'submitted_at'): continue
        db_col = col_map.get(key, key)
        sets.append(f"{db_col}=?")
        vals.append(val)
    
    if not sets: raise HTTPException(400, "Нет полей для обновления")
    vals.append(record_id)
    
    with get_db() as conn:
        r = conn.execute(f"SELECT id FROM {table} WHERE id=?", (record_id,)).fetchone()
        if not r: raise HTTPException(404, "Запись не найдена")
        conn.execute(f"UPDATE {table} SET {','.join(sets)} WHERE id=?", vals)
        
        # Пересчёт средней оценки для полевых выходов
        if table == "field_visits":
            row = conn.execute("SELECT haircut_quality,service_quality,additional_services_rating,cosmetics_rating,standards_rating FROM field_visits WHERE id=?", (record_id,)).fetchone()
            if row:
                avg = round((row[0]+row[1]+row[2]+row[3]+row[4])/5, 1)
                conn.execute("UPDATE field_visits SET average_rating=? WHERE id=?", (avg, record_id))
    
    return {"success": True, "message": "Запись обновлена"}

@app.delete("/record/{section}/{record_id}")
def delete_record(section: str, record_id: int):
    """Универсальное удаление записи по id"""
    cfg = SECTION_CONFIG.get(section)
    if not cfg: raise HTTPException(400, f"Неизвестная секция: {section}")
    with get_db() as conn:
        r = conn.execute(f"SELECT id FROM {cfg['table']} WHERE id=?", (record_id,)).fetchone()
        if not r: raise HTTPException(404, "Запись не найдена")
        conn.execute(f"DELETE FROM {cfg['table']} WHERE id=?", (record_id,))
    return {"success": True, "message": "Запись удалена"}

# ============= DASHBOARD =============
@app.get("/dashboard-summary/{branch_name}")
def get_dashboard_summary(branch_name: str):
    cm = current_month_ru()
    with get_db() as conn:
        if not conn.execute("SELECT id FROM branches WHERE name=?", (branch_name,)).fetchone():
            raise HTTPException(404, f"Филиал '{branch_name}' не найден")
        me = count_for_month(conn.execute("SELECT submitted_at FROM morning_events WHERE branch_name=?", (branch_name,)).fetchall(), 'submitted_at', cm)
        fv = count_for_month(conn.execute("SELECT submitted_at FROM field_visits WHERE branch_name=?", (branch_name,)).fetchall(), 'submitted_at', cm)
        oo = count_for_month(conn.execute("SELECT submitted_at FROM one_on_one WHERE branch_name=?", (branch_name,)).fetchall(), 'submitted_at', cm)
        mp = count_for_month(conn.execute("SELECT submitted_at FROM master_plans WHERE branch_name=?", (branch_name,)).fetchall(), 'submitted_at', cm)
        wm = count_for_month(conn.execute("SELECT submitted_at FROM weekly_metrics WHERE branch_name=?", (branch_name,)).fetchall(), 'submitted_at', cm)
        rv = sum_reviews_month(conn.execute("SELECT submitted_at, fact FROM reviews WHERE branch_name=?", (branch_name,)).fetchall(), cm)
        na = count_for_month(conn.execute("SELECT submitted_at FROM newbie_adaptation WHERE branch_name=?", (branch_name,)).fetchall(), 'submitted_at', cm)
    
    summary = {
        "morning_events": {"current":me,"goal":BRANCH_GOALS["morning_events"],"percentage":0,"label":"Утренние мероприятия"},
        "field_visits": {"current":fv,"goal":BRANCH_GOALS["field_visits"],"percentage":0,"label":"Полевые выходы"},
        "one_on_one": {"current":oo,"goal":BRANCH_GOALS["one_on_one"],"percentage":0,"label":"One-on-One"},
        "master_plans": {"current":mp,"goal":BRANCH_GOALS["master_plans"],"percentage":0,"label":"Планы мастеров"},
        "weekly_reports": {"current":wm,"goal":BRANCH_GOALS["weekly_reports"],"percentage":0,"label":"Еженедельные отчёты"},
        "reviews": {"current":rv,"goal":BRANCH_GOALS["reviews"],"percentage":0,"label":"Отзывы"},
        "new_employees": {"current":na,"goal":BRANCH_GOALS["new_employees"],"percentage":0,"label":"Новые сотрудники"},
    }
    for k in summary:
        if summary[k]["goal"] > 0: summary[k]["percentage"] = round((summary[k]["current"]/summary[k]["goal"])*100, 1)
    return {"success": True, "summary": summary}

# ============= CRUD ENDPOINTS =============
# --- Morning Events ---
@app.post("/morning-events/{branch_name}")
def submit_morning_events(branch_name: str, events: List[MorningEvent]):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as conn:
        for e in events:
            conn.execute("INSERT INTO morning_events (branch_name,submitted_at,date,week,event_type,participants,efficiency,comment) VALUES (?,?,?,?,?,?,?,?)",
                (branch_name,ts,e.date,e.week,e.event_type,e.participants,e.efficiency,e.comment or ""))
    return {"success": True, "message": f"Добавлено {len(events)} мероприятий"}

@app.get("/morning-events/{branch_name}")
def get_morning_events(branch_name: str):
    return get_section_data(branch_name, "morning-events")

# --- Field Visits ---
@app.post("/field-visits/{branch_name}")
def submit_field_visits(branch_name: str, visits: List[FieldVisit]):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as conn:
        for v in visits:
            avg = round((v.haircut_quality+v.service_quality+v.additional_services_rating+v.cosmetics_rating+v.standards_rating)/5, 1)
            conn.execute("INSERT INTO field_visits (branch_name,submitted_at,date,master_name,haircut_quality,service_quality,additional_services_comment,additional_services_rating,cosmetics_comment,cosmetics_rating,standards_comment,standards_rating,errors_comment,next_check_date,average_rating) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (branch_name,ts,v.date,v.master_name,v.haircut_quality,v.service_quality,v.additional_services_comment,v.additional_services_rating,v.cosmetics_comment,v.cosmetics_rating,v.standards_comment,v.standards_rating,v.errors_comment,v.next_check_date or "",avg))
    return {"success": True, "message": f"Добавлено {len(visits)} посещений"}

@app.get("/field-visits/{branch_name}")
def get_field_visits(branch_name: str):
    return get_section_data(branch_name, "field-visits")

# --- One-on-One ---
@app.post("/one-on-one/{branch_name}")
def submit_one_on_one(branch_name: str, meetings: List[OneOnOneMeeting]):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as conn:
        for m in meetings:
            conn.execute("INSERT INTO one_on_one (branch_name,submitted_at,date,master_name,goal,results,development_plan,indicator,next_meeting_date) VALUES (?,?,?,?,?,?,?,?,?)",
                (branch_name,ts,m.date,m.master_name,m.goal,m.results,m.development_plan,m.indicator,m.next_meeting_date or ""))
    return {"success": True, "message": f"Добавлено {len(meetings)} встреч"}

@app.get("/one-on-one/{branch_name}")
def get_one_on_one(branch_name: str):
    return get_section_data(branch_name, "one-on-one")

# --- Weekly Metrics ---
@app.post("/weekly-metrics/{branch_name}")
def submit_weekly_metrics(branch_name: str, metrics: List[WeeklyMetrics]):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as conn:
        for m in metrics:
            conn.execute("INSERT INTO weekly_metrics (branch_name,submitted_at,period,average_check_plan,average_check_fact,cosmetics_plan,cosmetics_fact,additional_services_plan,additional_services_fact) VALUES (?,?,?,?,?,?,?,?,?)",
                (branch_name,ts,m.period,m.average_check_plan,m.average_check_fact,m.cosmetics_plan,m.cosmetics_fact,m.additional_services_plan,m.additional_services_fact))
    return {"success": True, "message": f"Добавлено {len(metrics)} показателей"}

@app.get("/weekly-metrics/{branch_name}")
def get_weekly_metrics(branch_name: str):
    return get_section_data(branch_name, "weekly-metrics")

# --- Newbie Adaptation ---
@app.post("/newbie-adaptation/{branch_name}")
def submit_newbie_adaptation(branch_name: str, newbies: List[NewbieAdaptation]):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as conn:
        for n in newbies:
            conn.execute("INSERT INTO newbie_adaptation (branch_name,submitted_at,start_date,name,haircut_practice,service_standards,hygiene_sanitation,additional_services,cosmetics_sales,iclient_basics,status) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (branch_name,ts,n.start_date,n.name,n.haircut_practice,n.service_standards,n.hygiene_sanitation,n.additional_services,n.cosmetics_sales,n.iclient_basics,n.status))
    return {"success": True, "message": f"Добавлено {len(newbies)} записей"}

@app.get("/newbie-adaptation/{branch_name}")
def get_newbie_adaptation(branch_name: str):
    return get_section_data(branch_name, "newbie-adaptation")

# --- Master Plans ---
@app.post("/master-plans/{branch_name}")
def submit_master_plans(branch_name: str, plans: List[MasterPlan]):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as conn:
        for p in plans:
            conn.execute("INSERT INTO master_plans (branch_name,submitted_at,month,master_name,average_check_plan,average_check_fact,additional_services_plan,additional_services_fact,sales_plan,sales_fact,salary_plan,salary_fact) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (branch_name,ts,p.month,p.master_name,p.average_check_plan,p.average_check_fact,p.additional_services_plan,p.additional_services_fact,p.sales_plan,p.sales_fact,p.salary_plan,p.salary_fact))
    return {"success": True, "message": f"Добавлено {len(plans)} планов"}

@app.get("/master-plans/{branch_name}")
def get_master_plans(branch_name: str):
    return get_section_data(branch_name, "master-plans")

# --- Reviews ---
@app.post("/reviews/{branch_name}")
def submit_reviews(branch_name: str, reviews_list: List[Reviews]):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as conn:
        for r in reviews_list:
            conn.execute("INSERT INTO reviews (branch_name,submitted_at,week,manager_name,plan,fact,monthly_target) VALUES (?,?,?,?,?,?,?)",
                (branch_name,ts,r.week,r.manager_name,r.plan,r.fact,r.monthly_target))
    return {"success": True, "message": f"Добавлено {len(reviews_list)} отзывов"}

@app.get("/reviews/{branch_name}")
def get_reviews(branch_name: str):
    return get_section_data(branch_name, "reviews")

# --- Branch Summary ---
@app.post("/branch-summary/{branch_name}")
def generate_branch_summary(branch_name: str, summary: BranchSummary):
    with get_db() as conn:
        conn.execute("DELETE FROM branch_summaries WHERE branch_name=? AND month=?", (branch_name, summary.month))
        me = count_for_month(conn.execute("SELECT submitted_at FROM morning_events WHERE branch_name=?", (branch_name,)).fetchall(), 'submitted_at', summary.month)
        fv = count_for_month(conn.execute("SELECT submitted_at FROM field_visits WHERE branch_name=?", (branch_name,)).fetchall(), 'submitted_at', summary.month)
        oo = count_for_month(conn.execute("SELECT submitted_at FROM one_on_one WHERE branch_name=?", (branch_name,)).fetchall(), 'submitted_at', summary.month)
        mp = count_for_month(conn.execute("SELECT submitted_at FROM master_plans WHERE branch_name=?", (branch_name,)).fetchall(), 'submitted_at', summary.month)
        wm = count_for_month(conn.execute("SELECT submitted_at FROM weekly_metrics WHERE branch_name=?", (branch_name,)).fetchall(), 'submitted_at', summary.month)
        rv = sum_reviews_month(conn.execute("SELECT submitted_at, fact FROM reviews WHERE branch_name=?", (branch_name,)).fetchall(), summary.month)
        na = count_for_month(conn.execute("SELECT submitted_at FROM newbie_adaptation WHERE branch_name=?", (branch_name,)).fetchall(), 'submitted_at', summary.month)
        
        metrics = {"Утренние мероприятия":(me,BRANCH_GOALS["morning_events"]),"Полевые выходы":(fv,BRANCH_GOALS["field_visits"]),"One-on-One":(oo,BRANCH_GOALS["one_on_one"]),"Планы мастеров":(mp,BRANCH_GOALS["master_plans"]),"Еженедельные отчёты":(wm,BRANCH_GOALS["weekly_reports"]),"Отзывы":(rv,BRANCH_GOALS["reviews"]),"Новые сотрудники":(na,BRANCH_GOALS["new_employees"])}
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for name,(cur,goal) in metrics.items():
            pct = round((cur/goal)*100,1) if goal>0 else 0
            conn.execute("INSERT INTO branch_summaries (branch_name,submitted_at,manager,month,metric,current_value,goal_value,percentage) VALUES (?,?,?,?,?,?,?,?)",
                (branch_name,ts,summary.manager,summary.month,name,cur,goal,pct))
    return {"success": True, "message": "Отчёт создан"}

@app.get("/branch-summary/{branch_name}")
def get_branch_summary(branch_name: str):
    return get_section_data(branch_name, "branch-summary")

# ============= ADMIN: DASHBOARDS =============
@app.get("/admin/all-dashboards")
def admin_all_dashboards(period: str = Query("month")):
    start, end, label = get_period_dates(period)
    result = []
    with get_db() as conn:
        branches = conn.execute("SELECT name, manager_name FROM branches ORDER BY name").fetchall()
        for b in branches:
            bn = b['name']
            def cnt(table):
                rows = conn.execute(f"SELECT submitted_at FROM {table} WHERE branch_name=?", (bn,)).fetchall()
                return len([r for r in rows if parse_date_flexible(str(r['submitted_at'])) and start.date() <= parse_date_flexible(str(r['submitted_at'])).date() <= end.date()])
            def rv_sum():
                rows = conn.execute("SELECT submitted_at, fact FROM reviews WHERE branch_name=?", (bn,)).fetchall()
                return sum(int(r['fact'] or 0) for r in rows if parse_date_flexible(str(r['submitted_at'])) and start.date() <= parse_date_flexible(str(r['submitted_at'])).date() <= end.date())
            
            result.append({
                "branch_name": bn, "manager": b['manager_name'], "period_label": label,
                "morning_events": {"current": cnt("morning_events"), "goal": BRANCH_GOALS["morning_events"]},
                "field_visits": {"current": cnt("field_visits"), "goal": BRANCH_GOALS["field_visits"]},
                "one_on_one": {"current": cnt("one_on_one"), "goal": BRANCH_GOALS["one_on_one"]},
                "master_plans": {"current": cnt("master_plans"), "goal": BRANCH_GOALS["master_plans"]},
                "weekly_reports": {"current": cnt("weekly_metrics"), "goal": BRANCH_GOALS["weekly_reports"]},
                "reviews": {"current": rv_sum(), "goal": BRANCH_GOALS["reviews"]},
                "new_employees": {"current": cnt("newbie_adaptation"), "goal": BRANCH_GOALS["new_employees"]},
            })
    return {"success": True, "data": result, "period_label": label}

@app.get("/admin/branch-data/{branch_name}/{section}")
def admin_get_branch_data(branch_name: str, section: str, period: str = Query("all")):
    data = get_section_data(branch_name, section)
    if period != "all":
        start, end, label = get_period_dates(period)
        data["data"] = filter_by_period(data["data"], start, end)
        data["period_label"] = label
    return data

# ============= EMAIL =============
def build_multi_sheet_xlsx(sheets_data):
    wb = Workbook(); wb.remove(wb.active)
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfl = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for sn, recs in sheets_data.items():
        if not recs: continue
        ws = wb.create_sheet(title=sn[:31]); hdrs = list(recs[0].keys())
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(row=1, column=ci, value=h); c.font=hf; c.fill=hfl; c.alignment=ha; c.border=tb
        for ri, rec in enumerate(recs, 2):
            for ci, h in enumerate(hdrs, 1):
                v = rec.get(h, "")
                if isinstance(v, str):
                    try: v = float(v) if '.' in v else int(v)
                    except: pass
                c = ws.cell(row=ri, column=ci, value=v); c.border=tb
        for ci, h in enumerate(hdrs, 1):
            ml = max(len(str(h)), *[len(str(ws.cell(row=r, column=ci).value or "")) for r in range(2, len(recs)+2)])
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(ml+3, 50)
        ws.freeze_panes = "A2"
    if not wb.sheetnames: return b""
    out = io.BytesIO(); wb.save(out); return out.getvalue()

def send_email_with_attachments(to_email, subject, body_html, attachments):
    if not SMTP_USER or not SMTP_PASSWORD: raise HTTPException(500, "SMTP не настроен")
    if not to_email: raise HTTPException(500, "REPORT_EMAIL_TO не настроен")
    msg = MIMEMultipart(); msg['From']=SMTP_USER; msg['To']=to_email; msg['Subject']=subject
    msg.attach(MIMEText(body_html, 'html', 'utf-8'))
    for att in attachments:
        if att["content"]:
            part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            part.set_payload(att["content"]); encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{att["filename"]}"'); msg.attach(part)
    try:
        if SMTP_USE_SSL:
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as s: s.login(SMTP_USER, SMTP_PASSWORD); s.send_message(msg)
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s: s.starttls(); s.login(SMTP_USER, SMTP_PASSWORD); s.send_message(msg)
    except Exception as e: raise HTTPException(500, f"Ошибка отправки: {e}")

@app.post("/send-report/{branch_name}")
def send_report_email(branch_name: str, request: EmailReportRequest):
    start, end, label = get_period_dates(request.period_type, request.custom_date)
    sections_map = {"Утренние мероприятия":"morning-events","Полевые выходы":"field-visits","One-on-One":"one-on-one","Планы мастеров":"master-plans","Еженедельные показатели":"weekly-metrics","Отзывы":"reviews","Адаптация новичков":"newbie-adaptation","Итоговые отчеты":"branch-summary"}
    sheets_data = {}; total = 0
    for name, section in sections_map.items():
        recs = get_section_data(branch_name, section).get("data", [])
        if request.period_type != "all": recs = filter_by_period(recs, start, end)
        if recs: sheets_data[name] = recs; total += len(recs)
    if total == 0: return {"success": False, "message": f"Нет данных за: {label}"}
    xlsx = build_multi_sheet_xlsx(sheets_data)
    fn = f"Отчёт_{branch_name.replace(' ','_')}_{label.replace(' ','_').replace('.','_')}.xlsx"
    html = f"<html><body><h2>Отчёт: {branch_name}</h2><p>Период: {label}</p><p>{len(sheets_data)} вкладок, {total} записей</p></body></html>"
    send_email_with_attachments(REPORT_EMAIL_TO, f"Отчёт {branch_name} — {label}", html, [{"filename":fn,"content":xlsx}])
    return {"success": True, "message": f"Отправлен на {REPORT_EMAIL_TO}", "period": label, "sheets_count": len(sheets_data), "total_records": total}

@app.get("/email-config")
def get_email_config():
    return {"configured": bool(SMTP_USER and SMTP_PASSWORD and REPORT_EMAIL_TO),
        "smtp_host": SMTP_HOST, "smtp_user": (SMTP_USER[:3]+"***") if SMTP_USER else "",
        "report_email": (REPORT_EMAIL_TO[:3]+"***") if REPORT_EMAIL_TO else ""}
